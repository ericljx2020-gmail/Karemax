#!/usr/bin/env python3
"""
Excel Document Generator
Extracts information from export_details.xlsx and generates baoguandan.xlsx and fapiao_xiangdan.xlsx
while preserving the original formatting.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
from datetime import datetime
import shutil

def extract_export_details(file_path):
    """Extract structured data from export_details.xlsx"""
    print("Reading export details...")
    
    # Read the file without header to preserve structure
    df = pd.read_excel(file_path, header=None)
    
    # Extract basic information
    export_info = {}
    
    # Extract company and contact information
    export_info['company'] = 'KAREMAX INDUSTRIAL LIMITED'
    export_info['address'] = 'NO.217, NO.2 BUILDING, NO.15 SOUTH OF RIYING RD, FREE TRADE ZONE, SHANGHAI, CHINA'
    
    # Find key information from the header rows
    for idx, row in df.iterrows():
        if idx < 15:  # Header information is in first 15 rows
            if pd.notna(row[0]) and 'To:' in str(row[0]):
                export_info['buyer'] = str(row[1]) if pd.notna(row[1]) else ''
            elif pd.notna(row[3]) and 'Quote No.' in str(row[3]):
                export_info['quote_no'] = str(row[4]) if pd.notna(row[4]) else ''
            elif pd.notna(row[3]) and 'Date:' in str(row[3]):
                export_info['date'] = str(row[4]) if pd.notna(row[4]) else ''
    
    # Extract product information (starts from row 14)
    products = []
    
    # Find the data rows (after the header)
    data_start_row = None
    for idx, row in df.iterrows():
        if pd.notna(row[0]) and str(row[0]).strip() == '1':  # First product item
            data_start_row = idx
            break
    
    if data_start_row is not None:
        current_row = data_start_row
        while current_row < len(df) and pd.notna(df.iloc[current_row, 0]):
            row_data = df.iloc[current_row]
            
            # Skip rows that don't contain product data
            if str(row_data[0]).strip() in ['TOTAL:', '24MAO0423S'] or not str(row_data[0]).strip().isdigit():
                current_row += 1
                continue
                
            try:
                product = {
                    'item_no': str(row_data[0]).strip() if pd.notna(row_data[0]) else '',
                    'description': str(row_data[1]).strip() if pd.notna(row_data[1]) else '',
                    'qty': float(row_data[4]) if pd.notna(row_data[4]) and str(row_data[4]).replace('.', '').isdigit() else 0,
                    'unit': str(row_data[5]).strip() if pd.notna(row_data[5]) else '',
                    'unit_price': float(row_data[6]) if pd.notna(row_data[6]) and str(row_data[6]).replace('.', '').isdigit() else 0,
                    'amount': float(row_data[8]) if pd.notna(row_data[8]) and str(row_data[8]).replace('.', '').isdigit() else 0,
                    'cartons': float(row_data[10]) if pd.notna(row_data[10]) and str(row_data[10]).replace('.', '').isdigit() else 0,
                    'gross_weight': float(row_data[17]) if pd.notna(row_data[17]) and str(row_data[17]).replace('.', '').isdigit() else 0,
                    'net_weight': float(row_data[19]) if pd.notna(row_data[19]) and str(row_data[19]).replace('.', '').isdigit() else 0,
                }
                
                if product['item_no']:  # Only add if we have an item number
                    products.append(product)
                    
            except (ValueError, IndexError) as e:
                print(f"Warning: Could not parse row {current_row}: {e}")
            
            current_row += 1
    
    export_info['products'] = products
    
    # Calculate totals
    export_info['total_amount'] = sum(p['amount'] for p in products)
    export_info['total_cartons'] = sum(p['cartons'] for p in products)
    export_info['total_gross_weight'] = sum(p['gross_weight'] for p in products)
    export_info['total_net_weight'] = sum(p['net_weight'] for p in products)
    
    print(f"Extracted {len(products)} products")
    print(f"Total amount: ${export_info['total_amount']:,.2f}")
    
    return export_info

def safe_cell_update(ws, row, col, value):
    """Safely update a cell value, handling merged cells"""
    try:
        cell = ws.cell(row=row, column=col)
        if hasattr(cell, 'value'):
            cell.value = value
        return True
    except Exception as e:
        print(f"Warning: Could not update cell ({row}, {col}): {e}")
        return False

def generate_baoguandan(export_info, template_path, output_path):
    """Generate baoguandan.xlsx from template and export data"""
    print("Generating baoguandan...")
    
    # Remove output file if it exists to avoid permission issues
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except Exception as e:
            print(f"Warning: Could not remove existing file {output_path}: {e}")
    
    # First, copy the template file to preserve all formatting
    shutil.copy2(template_path, output_path)
    
    # Load the copied workbook
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Date (approximate position based on template structure)
    current_date = datetime.now().strftime('%Y.%m.%d')
    
    # Find and fill key fields - these coordinates may need adjustment based on actual template
    try:
        # Look for specific cells to update
        for row in range(1, 40):  # Search in reasonable range
            for col in range(1, 20):
                try:
                    cell = ws.cell(row=row, column=col)
                    cell_value = cell.value
                    if cell_value and isinstance(cell_value, str):
                        # Update specific fields based on content
                        if '境外收货人' in cell_value:
                            # Fill buyer information in next row
                            safe_cell_update(ws, row+1, col, export_info.get('buyer', ''))
                        elif '合同协议号' in cell_value:
                            # Fill quote number
                            safe_cell_update(ws, row, col+2, export_info.get('quote_no', ''))
                except Exception:
                    continue  # Skip problematic cells
    
        # Add product details in the appropriate section
        product_start_row = None
        for row in range(1, 40):
            try:
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and '项号' in str(cell_value):
                    product_start_row = row + 1
                    break
            except Exception:
                continue
        
        if product_start_row:
            for i, product in enumerate(export_info['products'][:10]):  # Limit to 10 products
                row_num = product_start_row + i
                if row_num <= 50:  # Reasonable limit
                    safe_cell_update(ws, row_num, 1, product['item_no'])
                    safe_cell_update(ws, row_num, 2, product['description'][:50])  # Truncate if too long
                    safe_cell_update(ws, row_num, 8, product['qty'])
                    safe_cell_update(ws, row_num, 10, f"${product['unit_price']:.2f}")
    
    except Exception as e:
        print(f"Warning: Could not update some fields in baoguandan: {e}")
    
    # Save the updated file
    wb.save(output_path)
    print(f"Baoguandan saved to: {output_path}")

def generate_fapiao_xiangdan(export_info, template_path, output_path):
    """Generate fapiao_xiangdan.xlsx from template and export data"""
    print("Generating fapiao xiangdan...")
    
    # Remove output file if it exists to avoid permission issues
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except Exception as e:
            print(f"Warning: Could not remove existing file {output_path}: {e}")
    
    # First, copy the template file to preserve all formatting
    shutil.copy2(template_path, output_path)
    
    # Load the copied workbook
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Update date
    current_date = datetime.now()
    
    # Find and update the date field
    for row in range(1, 35):
        for col in range(1, 12):
            try:
                cell = ws.cell(row=row, column=col)
                cell_value = cell.value
                if cell_value and isinstance(cell_value, str) and 'Date' in cell_value:
                    # Update the adjacent cell with current date
                    safe_cell_update(ws, row, col+1, current_date)
                elif isinstance(cell_value, datetime):
                    # Replace existing date
                    safe_cell_update(ws, row, col, current_date)
            except Exception:
                continue
    
    # Add product information
    # Find the product section or use a safe starting position
    product_start_row = 15  # Default safe position
    
    # Look for existing product data to determine insertion point
    for row in range(10, 35):
        try:
            for col in range(1, 12):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    if any(keyword in cell_value.lower() for keyword in ['description', 'item', 'product']):
                        product_start_row = row + 1
                        break
        except Exception:
            continue
        if product_start_row != 15:
            break
    
    # Add products to the invoice
    for i, product in enumerate(export_info['products'][:15]):  # Limit to 15 products
        row_num = product_start_row + i
        if row_num <= 50:  # Reasonable limit
            # Add product information safely
            safe_cell_update(ws, row_num, 1, f"{i+1}.")
            safe_cell_update(ws, row_num, 2, product['description'][:100])  # Limit length
            safe_cell_update(ws, row_num, 3, f"Qty: {product['qty']} {product['unit']}")
            safe_cell_update(ws, row_num, 4, f"Unit Price: ${product['unit_price']:.2f}")
            safe_cell_update(ws, row_num, 5, f"Amount: ${product['amount']:.2f}")
    
    # Add total
    total_row = product_start_row + len(export_info['products']) + 2
    if total_row <= 50:
        safe_cell_update(ws, total_row, 2, "TOTAL:")
        safe_cell_update(ws, total_row, 5, f"${export_info['total_amount']:.2f}")
        
        # Try to bold the total row
        try:
            for col in range(1, 6):
                cell = ws.cell(row=total_row, column=col)
                if hasattr(cell, 'font'):
                    cell.font = Font(bold=True)
        except Exception:
            pass  # If formatting fails, continue without it
    
    # Save the updated file
    wb.save(output_path)
    print(f"Fapiao xiangdan saved to: {output_path}")

def main():
    """Main processing function"""
    print("Excel Document Generator")
    print("=" * 50)
    
    # File paths
    export_details_file = 'export_details.xlsx'
    baoguandan_template = 'baoguandan.xlsx'
    fapiao_template = 'fapiao_xiangdan.xlsx'
    
    baoguandan_output = 'baoguandan_generated.xlsx'
    fapiao_output = 'fapiao_xiangdan_generated.xlsx'
    
    # Check if input files exist
    required_files = [export_details_file, baoguandan_template, fapiao_template]
    missing_files = [f for f in required_files if not os.path.exists(f)]
    
    if missing_files:
        print("Error: The following required files are missing:")
        for file in missing_files:
            print(f"  - {file}")
        print("\nPlease ensure all required files are in the current directory.")
        sys.exit(1)
    
    try:
        # Extract data from export details
        export_info = extract_export_details(export_details_file)
        
        if not export_info['products']:
            print("Warning: No products found in export details file")
            return
        
        # Generate output files
        generate_baoguandan(export_info, baoguandan_template, baoguandan_output)
        generate_fapiao_xiangdan(export_info, fapiao_template, fapiao_output)
        
        print("\n" + "=" * 50)
        print("Processing completed successfully!")
        print(f"Generated files:")
        print(f"  - {baoguandan_output}")
        print(f"  - {fapiao_output}")
        print(f"\nProcessed {len(export_info['products'])} products")
        print(f"Total value: ${export_info['total_amount']:,.2f}")
        
    except Exception as e:
        print(f"Error during processing: {e}")
        print("Please check the input files and try again.")
        sys.exit(1)

if __name__ == "__main__":
    main() 